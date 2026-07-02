"""
Exportação avançada de relatórios — PDF e Excel (Fase 5.4)

PDF: via reportlab (sem dependência de headless browser)
Excel: via openpyxl

Cada função retorna bytes prontos para StreamingResponse.
"""
from __future__ import annotations

import io
import logging
from datetime import datetime, timezone
from typing import Optional

logger = logging.getLogger(__name__)


# ── PDF ───────────────────────────────────────────────────────────────────────

def generate_os_pdf(os_list: list[dict], org_nome: str, periodo: str) -> bytes:
    """Gera PDF de lista de OS."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except ImportError:
        raise RuntimeError("reportlab não instalado. Execute: pip install reportlab")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    elements = []

    # Cabeçalho
    elements.append(Paragraph(f"<b>AURIX — Ordens de Serviço</b>", styles["Title"]))
    elements.append(Paragraph(f"{org_nome} · {periodo}", styles["Normal"]))
    elements.append(Spacer(1, 0.5*cm))

    # Tabela
    headers = ["#", "Equipamento", "Tipo", "Status", "Prioridade", "Criada em", "Técnico", "MTTR (min)"]
    data = [headers]
    for o in os_list:
        data.append([
            str(o.get("numero", "")),
            (o.get("equipamento_nome") or "")[:30],
            str(o.get("tipo", "")),
            str(o.get("status", "")),
            str(o.get("prioridade", "")),
            str(o.get("created_at", ""))[:10],
            (o.get("tecnico_nome") or "")[:20],
            str(o.get("tempo_reparo") or ""),
        ])

    col_widths = [1.2*cm, 5*cm, 2.5*cm, 3*cm, 2.5*cm, 2.5*cm, 3.5*cm, 2.5*cm]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1e293b")),
        ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
        ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0), (-1,-1), 8),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#f8fafc")]),
        ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#e2e8f0")),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("LEFTPADDING", (0,0), (-1,-1), 4),
        ("RIGHTPADDING", (0,0), (-1,-1), 4),
    ]))
    elements.append(table)

    elements.append(Spacer(1, 0.5*cm))
    elements.append(Paragraph(
        f"Gerado em {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')} UTC · AURIX v4.0",
        styles["Normal"],
    ))

    doc.build(elements)
    buf.seek(0)
    return buf.read()


def generate_kpi_pdf(kpis: dict, org_nome: str, periodo: str) -> bytes:
    """Gera relatório PDF de KPIs do dashboard."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib import colors
    except ImportError:
        raise RuntimeError("reportlab não instalado")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    el = []

    el.append(Paragraph("<b>AURIX — Relatório de KPIs de Manutenção</b>", styles["Title"]))
    el.append(Paragraph(f"{org_nome} · {periodo}", styles["Normal"]))
    el.append(Spacer(1, 1*cm))

    kpi_rows = [
        ["Indicador", "Valor"],
        ["OS no Mês",          str(kpis.get("total_os_mes", 0))],
        ["OS Abertas",         str(kpis.get("os_abertas", 0))],
        ["OS Atrasadas",       str(kpis.get("os_atrasadas", 0))],
        ["MTTR (h)",           f"{kpis.get('mttr', 0):.1f}"],
        ["MTBF (h)",           f"{kpis.get('mtbf', 0):.1f}"],
        ["Disponibilidade (%)", f"{kpis.get('disponibilidade', 0):.1f}"],
        ["Custo Total (R$)",   f"R$ {kpis.get('custo_total_mes', 0):,.2f}"],
    ]

    t = Table(kpi_rows, colWidths=[8*cm, 6*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1e293b")),
        ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
        ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0), (-1,-1), 11),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#f8fafc")]),
        ("GRID", (0,0), (-1,-1), 0.5, colors.lightgrey),
        ("LEFTPADDING", (0,0), (-1,-1), 8),
        ("TOPPADDING",  (0,0), (-1,-1), 6),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
    ]))
    el.append(t)

    doc.build(el)
    buf.seek(0)
    return buf.read()


# ── Excel ─────────────────────────────────────────────────────────────────────

def generate_os_excel(os_list: list[dict], org_nome: str, periodo: str) -> bytes:
    """Gera Excel (.xlsx) com lista de OS."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise RuntimeError("openpyxl não instalado. Execute: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ordens de Serviço"

    # Estilos
    header_fill = PatternFill("solid", fgColor="1e293b")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        bottom=Side(style="thin", color="e2e8f0"),
        right=Side(style="thin", color="e2e8f0"),
    )

    # Metadados
    ws["A1"] = "AURIX — Ordens de Serviço"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"{org_nome} · {periodo}"
    ws["A3"] = f"Gerado em: {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')} UTC"

    # Cabeçalho
    headers = ["#", "Equipamento", "Tipo", "Status", "Prioridade",
               "Criada em", "Técnico", "Solução", "MTTR (min)", "Custo (R$)"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    # Dados
    alt_fill = PatternFill("solid", fgColor="f8fafc")
    for row_idx, o in enumerate(os_list, start=6):
        fill = alt_fill if row_idx % 2 == 0 else None
        row_data = [
            o.get("numero"),
            o.get("equipamento_nome", ""),
            o.get("tipo", ""),
            o.get("status", ""),
            o.get("prioridade", ""),
            str(o.get("created_at", ""))[:19],
            o.get("tecnico_nome", ""),
            (o.get("solucao") or "")[:100],
            o.get("tempo_reparo"),
            o.get("custo_parada"),
        ]
        for col, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = border
            if fill:
                cell.fill = fill

    # Auto-width
    col_widths = [8, 30, 15, 18, 14, 20, 22, 40, 14, 14]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
